# -*- coding:utf-8 -*-
import openpyxl
import datetime
import re
import time
import ctypes
import sys

#定义变量
current_vcomp='component4.xlsx'
release_filename='1100版本发布.txt'
current_line=""
models_dir={}
now=time.strftime("%Y%m%d %X",time.localtime())[0:8]

#读取exce
#active : 获取活跃的Worksheet
#read_only : 是否只读模式打开Excel文档
#encoding : 文档的字符编码
#properties : 文档的元数据，如标题，创建者，创建日期等
#worksheets : 以列表的形式返回所有的sheet
#获取sheet_master B1 的值
#B1=sheet_master2['B1'].value
#sheet_master2['C1']=datetime.datetime.now()
#wb.save('example.xlsx')
#sheet_master2.max_row 
#
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
STD_ERROR_HANDLE = -12
#定义字体颜色
FOREGROUND_BLACK = 0x00 # black.
FOREGROUND_DARKBLUE = 0x01 # dark blue.
FOREGROUND_DARKGREEN = 0x02 # dark green.
FOREGROUND_DARKSKYBLUE = 0x03 # dark skyblue.
FOREGROUND_DARKRED = 0x04 # dark red.
FOREGROUND_DARKPINK = 0x05 # dark pink.
FOREGROUND_DARKYELLOW = 0x06 # dark yellow.
FOREGROUND_DARKWHITE = 0x07 # dark white.
FOREGROUND_DARKGRAY = 0x08 # dark gray.
FOREGROUND_BLUE = 0x09 # blue.
FOREGROUND_GREEN = 0x0a # green.
FOREGROUND_SKYBLUE = 0x0b # skyblue.
FOREGROUND_RED = 0x0c # red.
FOREGROUND_PINK = 0x0d # pink.
FOREGROUND_YELLOW = 0x0e # yellow.
FOREGROUND_WHITE = 0x0f # white.


# Windows CMD命令行 背景颜色定义 background colors
BACKGROUND_BLUE = 0x10 # dark blue.
BACKGROUND_GREEN = 0x20 # dark green.
BACKGROUND_DARKSKYBLUE = 0x30 # dark skyblue.
BACKGROUND_DARKRED = 0x40 # dark red.
BACKGROUND_DARKPINK = 0x50 # dark pink.
BACKGROUND_DARKYELLOW = 0x60 # dark yellow.
BACKGROUND_DARKWHITE = 0x70 # dark white.
BACKGROUND_DARKGRAY = 0x80 # dark gray.
BACKGROUND_BLUE = 0x90 # blue.
BACKGROUND_GREEN = 0xa0 # green.
BACKGROUND_SKYBLUE = 0xb0 # skyblue.
BACKGROUND_RED = 0xc0 # red.
BACKGROUND_PINK = 0xd0 # pink.
BACKGROUND_YELLOW = 0xe0 # yellow.
BACKGROUND_WHITE = 0xf0 # white.
#get handle
# get handle
std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)

def set_cmd_text_color(color, handle=std_out_handle):
    Bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
    return Bool

#reset white
def resetColor():
    set_cmd_text_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE)

def printGreen(mess):
    set_cmd_text_color(FOREGROUND_GREEN)
    sys.stdout.write(mess)
    resetColor()
def printRed(mess):
    set_cmd_text_color(FOREGROUND_RED)
    sys.stdout.write(mess)
    resetColor()


def get_update_models(linecontent):
    for i in linecontent:
         pass

def get_current_num(filename):
    f=open(filename,'r', encoding='UTF-8')
    for num,value in enumerate(f):
        if value.strip() != "" and value.strip()[0:8] == now :
            return num  
    f.close()            
            
current_line=get_current_num(release_filename)
ff=open(release_filename,'r', encoding='UTF-8')
conent=ff.readlines()[current_line+1:]
for i in conent:
    if i.strip() != "" :
        t=re.split(r" +",i)
        models_dir[t[0]]={}
        models_dir[t[0]]['Author']=t[1]
        models_dir[t[0]]['oldtonew']= t[2][t[2].find('[')+1:t[2].find(']')]
 	
def change_component(component):
    wb=openpyxl.load_workbook(component)
    sheet_master2 = wb.worksheets[9]
    max_row=sheet_master2.max_row
    for n in range(1,max_row):
        if sheet_master2['B'+str(n)].value.lower()  in models_dir.keys():
            old_version=sheet_master2['C'+str(n)].value
            v_old_version=models_dir[sheet_master2['B'+str(n)].value.lower()]['oldtonew'].split('->')[0]
            v_new_version=models_dir[sheet_master2['B'+str(n)].value.lower()]['oldtonew'].split('->')[1]
            if old_version == v_old_version:
                sheet_master2['C'+str(n)]=v_new_version
                wb.save(component)
                printGreen (u"%s%s %s->%s\t%s\n"% (sheet_master2['B'+str(n)].value,\
                " "*(26-len(sheet_master2['B'+str(n)].value)),\
                old_version,v_new_version,\
                models_dir[sheet_master2['B'+str(n)].value.lower()]['Author']))
            else :
                sheet_master2['C'+str(n)]=v_new_version
                wb.save(component)            
                printRed (u"%s%s %s->%s\t%s\n"% (sheet_master2['B'+str(n)].value,\
                " "*(26-len(sheet_master2['B'+str(n)].value)),\
                old_version,v_new_version,\
                models_dir[sheet_master2['B'+str(n)].value.lower()]['Author']))
if __name__ == '__main__':
    change_component(current_vcomp) 
    print ("{0}".format(models_dir))
    print
             
    